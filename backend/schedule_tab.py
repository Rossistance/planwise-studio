"""Schedules from spreadsheets — Smartsheet and MS Project exports, or a CSV.

Both tools export the same handful of ideas under different labels: what
Project calls "Task Name" Smartsheet calls "Task Name" or "Primary Column",
"Finish" is sometimes "End Date", "% Complete" is sometimes "Progress". So the
mapping is by SYNONYM against the header row rather than by position, and a
column that isn't recognised is reported rather than dropped in silence.

Nothing here guesses at a value. A cell that won't parse as a date leaves the
date empty and raises a warning; it never becomes today, or the row above.
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any

from .schedule import ScheduleError, parse_date_text, parse_duration_text

# Header synonyms, lower-cased and stripped of punctuation. Order matters only
# in that the first match wins, so put the unambiguous ones first.
COLUMN_SYNONYMS: dict[str, tuple[str, ...]] = {
    "external_id": ("id", "task id", "unique id", "uid", "row id", "activity id",
                    "task number", "no", "#"),
    "wbs": ("wbs", "outline number", "wbs code", "outline"),
    "name": ("task name", "name", "title", "primary column", "task",
             "activity name", "description", "activity"),
    "duration_days": ("duration", "dur", "days"),
    "start": ("start", "start date", "planned start", "early start",
              "scheduled start", "begin"),
    "finish": ("finish", "finish date", "end", "end date", "planned finish",
               "early finish", "scheduled finish", "due date"),
    "percent_complete": ("% complete", "percent complete", "pct complete",
                         "progress", "complete", "% done"),
    "predecessors": ("predecessors", "predecessor", "depends on", "dependency",
                     "dependencies", "preds"),
    "outline_level": ("outline level", "level", "indent"),
    "notes": ("notes", "comments", "note", "remarks"),
}

# Columns worth naming when they exist but aren't imported, so nobody assumes
# they came across.
NOTED_COLUMNS = ("assigned to", "resource names", "resources", "owner", "cost",
                 "work", "baseline start", "baseline finish", "constraint type",
                 "constraint date", "deadline", "actual start", "actual finish")


def _norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9% ]+", " ", str(s or "").strip().lower()).strip()


def map_headers(headers: list[Any]) -> tuple[dict[str, int], list[str]]:
    """Header row -> {field: column index}, plus the headers left over.

    Exact synonym match first, then a contained match, so "Task Start Date"
    still finds `start` without "Start" also claiming "Baseline Start".
    """
    norm = [_norm(h) for h in headers]
    used: set[int] = set()
    mapping: dict[str, int] = {}

    for field, names in COLUMN_SYNONYMS.items():
        for i, h in enumerate(norm):
            if i in used or not h:
                continue
            if h in names:
                mapping[field] = i
                used.add(i)
                break
    for field, names in COLUMN_SYNONYMS.items():
        if field in mapping:
            continue
        for i, h in enumerate(norm):
            if i in used or not h:
                continue
            if any(n in h for n in names):
                mapping[field] = i
                used.add(i)
                break

    leftover = [str(headers[i]) for i, h in enumerate(norm)
                if h and i not in used]
    return mapping, leftover


def _rows_from_csv(data: bytes) -> list[list[Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel                      # a single column is still valid
    return [list(r) for r in csv.reader(io.StringIO(text), dialect)]


def _rows_from_xlsx(data: bytes) -> list[list[Any]]:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl raises many shapes
        raise ScheduleError(f"That spreadsheet could not be opened: {exc}") from exc
    ws = wb[wb.sheetnames[0]]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _find_header_row(rows: list[list[Any]]) -> int:
    """The first row that maps to a name and at least one date or duration.

    Exports routinely carry a title row or two above the real header, and
    taking row 0 on faith turns "Project: Siemens Wendell" into the column
    names and the whole file into nothing.
    """
    for i, row in enumerate(rows[:25]):
        mapping, _ = map_headers(row)
        if "name" in mapping and ({"start", "finish", "duration_days"} & set(mapping)):
            return i
    return -1


def parse_table(data: bytes, filename: str) -> dict[str, Any]:
    """Spreadsheet or CSV -> {tasks, links, warnings}.

    `links` is always empty: a table states its dependencies in a Predecessors
    column, which travels as text on the task and becomes real links when the
    import is committed — there is nothing to infer and nothing to confirm.
    """
    name = (filename or "").lower()
    rows = _rows_from_csv(data) if name.endswith(".csv") else _rows_from_xlsx(data)
    rows = [r for r in rows if any(str(c).strip() for c in r if c is not None)]
    if not rows:
        raise ScheduleError("That file has no rows in it.")

    hdr = _find_header_row(rows)
    if hdr < 0:
        raise ScheduleError(
            "No schedule columns were recognised. The sheet needs a header row "
            "with at least a task name and a start, finish or duration column.")

    mapping, leftover = map_headers(rows[hdr])
    warnings: list[str] = []
    noted = [h for h in leftover if _norm(h) in NOTED_COLUMNS]
    if noted:
        warnings.append(
            "These columns were found but not imported, because PlanWise has "
            f"nowhere to put them yet: {', '.join(noted)}.")
    other = [h for h in leftover if _norm(h) not in NOTED_COLUMNS]
    if other:
        warnings.append(f"Columns ignored: {', '.join(other[:8])}"
                        + (" …" if len(other) > 8 else "") + ".")

    def cell(row: list[Any], field: str) -> Any:
        i = mapping.get(field)
        if i is None or i >= len(row):
            return None
        v = row[i]
        return v.strip() if isinstance(v, str) else v

    tasks: list[dict[str, Any]] = []
    for n, row in enumerate(rows[hdr + 1:], start=hdr + 2):
        raw_name = cell(row, "name")
        if raw_name is None or not str(raw_name).strip():
            continue
        start, start_ok = parse_date_text(cell(row, "start"))
        finish, finish_ok = parse_date_text(cell(row, "finish"))
        if not start_ok or not finish_ok:
            warnings.append(f"Row {n}: a date names a weekday that doesn't match it.")
        if cell(row, "start") and not start:
            warnings.append(f"Row {n} ({str(raw_name)[:30]}): start date "
                            f"'{cell(row, 'start')}' could not be read; left empty.")
        dur, unit = parse_duration_text(cell(row, "duration_days"))

        pct = cell(row, "percent_complete")
        if isinstance(pct, str):
            pct = pct.strip().rstrip("%") or None
        try:
            pct = float(pct) if pct is not None else 0.0
        except (TypeError, ValueError):
            pct = 0.0
        if 0 < pct <= 1 and isinstance(cell(row, "percent_complete"), float):
            pct *= 100                       # a spreadsheet percentage is a fraction

        try:
            level = int(cell(row, "outline_level") or 1)
        except (TypeError, ValueError):
            level = 1

        wbs = cell(row, "wbs")
        wbs = str(wbs).strip() if wbs not in (None, "") else None
        if wbs and mapping.get("outline_level") is None:
            level = len(wbs.split("."))

        ext = cell(row, "external_id")
        preds = cell(row, "predecessors")
        tasks.append({
            "external_id": str(ext).strip() if ext not in (None, "") else str(n),
            "wbs": wbs,
            "name": str(raw_name).strip(),
            "start": start,
            "finish": finish,
            "duration_days": dur,
            "duration_unit": unit,
            "percent_complete": pct,
            "outline_level": max(1, level),
            "is_milestone": 1 if dur == 0 else 0,
            "is_summary": 0,
            "predecessors": str(preds).strip() if preds not in (None, "") else None,
            "notes": (str(cell(row, "notes")).strip()
                      if cell(row, "notes") not in (None, "") else None),
            "sort_order": len(tasks),
        })

    if not tasks:
        raise ScheduleError("The header row was found but no task rows followed it.")

    # Summary rows are structural: a row whose successor is deeper.
    for i, t in enumerate(tasks):
        nxt = tasks[i + 1] if i + 1 < len(tasks) else None
        t["is_summary"] = 1 if nxt and nxt["outline_level"] > t["outline_level"] else 0
        if t["is_summary"]:
            t["is_milestone"] = 0

    return {"tasks": tasks, "links": [], "warnings": warnings}
