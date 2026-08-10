"""Read the shared Vista extract.

    Power BI  BIQLDataModelVista   (refreshed daily ~3:37 AM)
          |
          v   Power Automate flow (cloud, no PC involved)
    Vista Model 2026 - Data.xlsx   values-only tables + Meta.as_of
          |
          v   OneDrive sync -> this module (read-only, no sign-in)
    Snapshot

Four sheets, verified against the live workbook on 2026-08-08:

    'Pivot Data'        one row per job (9,318 jobs), 16 columns
    'WRH Phase Detail'  one row per Job x Phase x Cost Type (134,983 rows)
    'WRH Job Status'    job -> status / contract type
    'Meta'              as_of stamp written by the flow

Two rules this module holds to, both inherited from hard-won lessons in the
2.0 line:

* **None is not zero.** A blank cell means Vista reported nothing; zero means
  Vista reported zero. Collapsing them produced dashboards that looked
  confident about numbers nobody had. Blanks stay None all the way to the UI.
* **Cost types are data, not a constant.** The old build hardcoded nine cost
  types and dropped anything else. Here they come from whatever
  'Cost Type Desc' actually contains for the job in front of you.
"""
from __future__ import annotations

import threading
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from typing import Any

from . import config

# --- sheet + column contract ------------------------------------------------
# Named so a schema drift fails loudly at load with a useful message, rather
# than silently yielding empty columns.

SHEET_JOBS = "Pivot Data"
SHEET_PHASES = "WRH Phase Detail"
SHEET_STATUS = "WRH Job Status"
SHEET_CONTRACT_AR = "WRH Contract AR"  # schema v2
SHEET_META = "Meta"

JOB_COLUMNS = {
    "Row Labels": "job_label",
    "Job Number": "job_number",
    "Job Name": "job_name",
    "Current Contract Amt": "current_contract",
    "Original Contract Amt": "original_contract",
    "Change Order Revenue": "change_order_revenue",
    "Actual Billed": "actual_billed",
    "Actual Cost": "actual_cost",
    "Projected Cost": "projected_cost",
    "Current Estimate Costs": "current_estimate",
    "Earned Revenue - JTD": "earned_revenue",
    "Actual Cost - JTD Labor": "labor_cost",
    "Actual Hours - JTD": "labor_hours",
    "Actual % Complete Estimated - JTD": "pct_complete",
    "Financial Status": "financial_status",
    "Contract Size Band": "size_band",
}

# Schema v2 job columns. Kept separate from JOB_COLUMNS so a v1 workbook
# (e.g. a machine whose OneDrive hasn't synced the regenerated file yet)
# still loads — these fields just come back None, which the UI already
# renders honestly as "not reported".
JOB_COLUMNS_V2 = {
    "Actual Cost - MTD": "mtd_cost",
    "Actual Hours - MTD": "mtd_hours",
    "Actual Billed - MTD": "mtd_billed",
    "AP Unapproved Invoice Amt": "unapproved_ap",
}

PHASE_COLUMNS_V2 = {
    "Actual Cost - MTD": "mtd_cost",
}

# Contract numbers match job numbers; verified to the penny against the
# Vista Crystal export for 24-003 (billed 5,859,364.00 / received
# 5,273,427.62 / retainage 585,936.38).
CONTRACT_AR_COLUMNS = {
    "Contract": "contract",
    "Contract Status": "contract_status",
    "Billed Amt": "billed",
    "Received Amt": "collected",
    "Current Retain Amt": "retainage",
}

PHASE_COLUMNS = {
    "Job and Desc": "job_label",
    "Phase and Desc": "phase",
    "Cost Type Desc": "cost_type",
    "Actual Cost": "actual_cost",
    "Current Estimate Costs": "current_estimate",
    "Projected Cost": "projected_cost",
    "Actual Hours - JTD": "hours_units",
    "Remaining Cost - JTD": "remaining_cost",
}

STATUS_COLUMNS = {
    "Job and Desc": "job_label",
    "Job Status": "job_status",
    "Contract Type": "contract_type",
}


class VistaUnavailable(RuntimeError):
    """The workbook is missing, unreadable, or not the shape we expect."""


@dataclass
class Snapshot:
    as_of: datetime | None
    source_path: str
    schema_version: int = 1
    jobs: dict[str, dict[str, Any]] = field(default_factory=dict)
    phases: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    # job number -> {billed, collected, retainage, contract_status}; empty on
    # a v1 workbook.
    contract_ar: dict[str, dict[str, Any]] = field(default_factory=dict)

    @property
    def is_stale(self) -> bool:
        if self.as_of is None:
            return True
        age = datetime.now(timezone.utc) - self.as_of.astimezone(timezone.utc)
        return age > timedelta(hours=config.STALE_AFTER_HOURS)

    @property
    def age_hours(self) -> float | None:
        if self.as_of is None:
            return None
        age = datetime.now(timezone.utc) - self.as_of.astimezone(timezone.utc)
        return round(age.total_seconds() / 3600, 1)


# --- cache ------------------------------------------------------------------
# 135k phase rows takes ~8s to parse. Hold it in memory and re-read only when
# OneDrive lands a new file (mtime + size change).

_lock = threading.Lock()
_cached: tuple[tuple[float, int], Snapshot] | None = None


def _num(v: Any) -> float | None:
    """Vista numerics. Blank stays blank; zero stays zero."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except ValueError:
        return None


def _text(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _header_index(header: tuple, wanted: dict[str, str], sheet: str,
                  optional: dict[str, str] | None = None) -> dict[str, int]:
    """Map our field names to column positions, by caption.

    `wanted` captions are required — a schema drift fails loudly. `optional`
    captions (schema-v2 additions) map when present and are skipped when the
    workbook predates them.
    """
    seen = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    missing = [cap for cap in wanted if cap not in seen]
    if missing:
        raise VistaUnavailable(
            f"Sheet '{sheet}' is missing expected column(s): {', '.join(missing)}. "
            f"Found: {', '.join(sorted(seen))}"
        )
    idx = {field: seen[cap] for cap, field in wanted.items()}
    for cap, field in (optional or {}).items():
        if cap in seen:
            idx[field] = seen[cap]
    return idx


def load(force: bool = False) -> Snapshot:
    """Return the current snapshot, re-reading only when the file changed."""
    global _cached

    path = config.vista_workbook()
    if path is None:
        raise VistaUnavailable(
            "The Vista workbook was not found. It syncs to "
            "'OneDrive - <tenant>/Company Share - Documents/Vista Model 2026- RH/"
            "Vista Model 2026 - Data.xlsx'. If OneDrive is still syncing, wait; "
            "otherwise set PLANWISE_VISTA_WORKBOOK to its path."
        )

    stat = path.stat()
    stamp = (stat.st_mtime, stat.st_size)

    with _lock:
        if not force and _cached is not None and _cached[0] == stamp:
            return _cached[1]
        snap = _read(path)
        _cached = (stamp, snap)
        return snap


def _read(path) -> Snapshot:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise VistaUnavailable("openpyxl is required to read the Vista workbook") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for required in (SHEET_JOBS, SHEET_PHASES, SHEET_STATUS):
            if required not in wb.sheetnames:
                raise VistaUnavailable(
                    f"'{path.name}' has no '{required}' sheet "
                    f"(found: {', '.join(wb.sheetnames)}). This does not look like "
                    "the Vista Model 2026 extract."
                )

        as_of = _read_meta_as_of(wb)
        schema_version = _read_meta_schema_version(wb)
        jobs = _read_jobs(wb)
        _apply_status(wb, jobs)
        phases = _read_phases(wb)
        contract_ar = _read_contract_ar(wb)
    finally:
        wb.close()

    if not jobs:
        raise VistaUnavailable(
            f"'{path.name}' parsed to zero jobs. The file is present but empty or "
            "mid-write; nothing has been loaded."
        )

    return Snapshot(as_of=as_of, source_path=str(path), schema_version=schema_version,
                    jobs=jobs, phases=phases, contract_ar=contract_ar)


def _read_meta_as_of(wb) -> datetime | None:
    if SHEET_META not in wb.sheetnames:
        return None
    for row in wb[SHEET_META].iter_rows(values_only=True):
        if row and _text(row[0]) == "as_of":
            raw = row[1]
            if isinstance(raw, datetime):
                return raw
            try:
                return datetime.fromisoformat(str(raw))
            except (TypeError, ValueError):
                return None
    return None


def _read_meta_schema_version(wb) -> int:
    if SHEET_META not in wb.sheetnames:
        return 1
    for row in wb[SHEET_META].iter_rows(values_only=True):
        if row and _text(row[0]) == "schema_version":
            try:
                return int(str(row[1]))
            except (TypeError, ValueError):
                return 1
    return 1


def _read_jobs(wb) -> dict[str, dict[str, Any]]:
    ws = wb[SHEET_JOBS]
    rows = ws.iter_rows(values_only=True)
    idx = _header_index(next(rows), JOB_COLUMNS, SHEET_JOBS, optional=JOB_COLUMNS_V2)

    numeric = {
        "current_contract", "original_contract", "change_order_revenue",
        "actual_billed", "actual_cost", "projected_cost", "current_estimate",
        "earned_revenue", "labor_cost", "labor_hours", "pct_complete",
        "mtd_cost", "mtd_hours", "mtd_billed", "unapproved_ap",
    }

    jobs: dict[str, dict[str, Any]] = {}
    for row in rows:
        number = _text(row[idx["job_number"]])
        if not number:
            continue
        # v2 fields default to None so a v1 workbook yields "not reported",
        # never a silent 0.
        rec: dict[str, Any] = {f: None for f in JOB_COLUMNS_V2.values()}
        for fieldname, col in idx.items():
            rec[fieldname] = _num(row[col]) if fieldname in numeric else _text(row[col])
        rec["job_status"] = None
        rec["contract_type"] = None
        jobs[number] = rec
    return jobs


def _read_contract_ar(wb) -> dict[str, dict[str, Any]]:
    """Schema v2 sheet; absent on a v1 workbook (empty dict, not an error)."""
    if SHEET_CONTRACT_AR not in wb.sheetnames:
        return {}
    ws = wb[SHEET_CONTRACT_AR]
    rows = ws.iter_rows(values_only=True)
    idx = _header_index(next(rows), CONTRACT_AR_COLUMNS, SHEET_CONTRACT_AR)

    numeric = {"billed", "collected", "retainage"}
    out: dict[str, dict[str, Any]] = {}
    for row in rows:
        contract = _text(row[idx["contract"]])
        if not contract:
            continue
        out[contract] = {
            f: (_num(row[c]) if f in numeric else _text(row[c]))
            for f, c in idx.items()
        }
    return out


def _apply_status(wb, jobs: dict[str, dict[str, Any]]) -> None:
    """Fold job status / contract type onto the job records, keyed by label."""
    ws = wb[SHEET_STATUS]
    rows = ws.iter_rows(values_only=True)
    idx = _header_index(next(rows), STATUS_COLUMNS, SHEET_STATUS)

    by_label = {rec["job_label"]: rec for rec in jobs.values() if rec.get("job_label")}
    for row in rows:
        rec = by_label.get(_text(row[idx["job_label"]]))
        if rec is None:
            continue
        rec["job_status"] = _text(row[idx["job_status"]])
        rec["contract_type"] = _text(row[idx["contract_type"]])


def _read_phases(wb) -> dict[str, list[dict[str, Any]]]:
    ws = wb[SHEET_PHASES]
    rows = ws.iter_rows(values_only=True)
    idx = _header_index(next(rows), PHASE_COLUMNS, SHEET_PHASES,
                        optional=PHASE_COLUMNS_V2)

    numeric = {
        "actual_cost", "current_estimate", "projected_cost",
        "hours_units", "remaining_cost", "mtd_cost",
    }

    out: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        label = _text(row[idx["job_label"]])
        if not label:
            continue
        rec = {
            f: (_num(row[c]) if f in numeric else _text(row[c]))
            for f, c in idx.items()
            if f != "job_label"
        }
        out.setdefault(label, []).append(rec)
    return out


# --- derived views ----------------------------------------------------------

def job_numbers(snap: Snapshot) -> list[str]:
    """Job numbers sorted naturally: 1, 2, 2-017, 3, ... 10, 24-003."""
    def key(n: str):
        parts = n.split("-")
        head = parts[0]
        return (0, int(head), n) if head.isdigit() else (1, 0, n)
    return sorted(snap.jobs, key=key)


def phases_for(snap: Snapshot, job_number: str) -> list[dict[str, Any]]:
    job = snap.jobs.get(job_number)
    if not job or not job.get("job_label"):
        return []
    return snap.phases.get(job["job_label"], [])


def cost_types_for(snap: Snapshot, job_number: str) -> list[dict[str, Any]]:
    """Roll the job's phase lines up to cost type.

    The cost types are whatever this job actually has — no fixed taxonomy, no
    zero-filled rows for types the job never used.
    """
    buckets: dict[str, dict[str, Any]] = {}
    for row in phases_for(snap, job_number):
        name = row.get("cost_type") or "Unclassified"
        b = buckets.setdefault(name, {
            "cost_type": name,
            "actual_cost": None,
            "current_estimate": None,
            "projected_cost": None,
            "hours_units": None,
            "mtd_cost": None,
            "phase_count": 0,
        })
        b["phase_count"] += 1
        for f in ("actual_cost", "current_estimate", "projected_cost",
                  "hours_units", "mtd_cost"):
            v = row.get(f)
            if v is not None:
                b[f] = (b[f] or 0.0) + v

    for b in buckets.values():
        est, act, proj = b["current_estimate"], b["actual_cost"], b["projected_cost"]
        # Variance is estimate vs money actually spent — budget remaining on the
        # line. Deliberately NOT estimate vs Projected Cost: Vista's projection
        # can sit below actual cost on a stale line (job 24-003 Equipment reads
        # est 67,379 / actual 165,065 / projected 55,379), and estimate-minus-
        # projected reports that as a $12,000 surplus on a line $98K overspent.
        b["variance"] = None if (est is None or act is None) else est - act
        b["pct_complete"] = None if (not proj or act is None) else act / proj

    return sorted(buckets.values(), key=lambda b: -(b["actual_cost"] or 0))
