"""Seed an ISOLATED PlanWise instance with synthetic data for development.

The 2.0 rebuild (and the release-verification checklist) needs an instance
where every page has something on it and nothing is real: mutations during
development must never land in ~/.planwise, and screenshots must be safe to
share. All names, jobs and figures here are invented — the same fictional
cast the .stitch design captures used (job 26-101 Northwind etc.).

Usage:
    .venv/Scripts/python.exe scripts/seed-dev.py --data-dir <dir>

Writes: a synthetic v2 Vista workbook, an approved admin account
(dev@planwise.local / planwise-dev), a second approved teammate, and per-job
registers: COs (customer + sub, one drafted-unsent), POs + invoices (one
issued from a sub CO, one exposure left open), RFIs/submittals (a draft, a
sent, one with an unconfirmed reply), a schedule with all four link types,
look-ahead rows with areas, and a two-extract vista_history so the forecast
chart has a line to draw.
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def build_workbook(path: Path, as_of: date) -> None:
    from openpyxl import Workbook
    wb = Workbook()

    jobs = [
        # num, name, orig, co_rev, billed, actual, projected, estimate, status
        ("26-101", "Northwind Substation Upgrade", 5940000, 542910, 4196540, 3884215, 5318900, 5318700, "On track"),
        ("26-114", "Cascade Water Treatment Expansion", 2450000, 0, 980000, 1042000, 2380000, 2400000, "On track"),
        ("25-088", "Brightline Logistics Center", 8900000, 231000, 8102000, 7455000, 8541000, 8600000, "Overbilled"),
        ("26-127", "Harbor Point Data Hall", 12400000, 0, 1240000, 1868000, 12100000, 12200000, "On track"),
    ]

    ws = wb.active
    ws.title = "Pivot Data"
    ws.append(["Row Labels", "Job Number", "Job Name", "Current Contract Amt",
               "Original Contract Amt", "Change Order Revenue", "Actual Billed",
               "Actual Cost", "Projected Cost", "Current Estimate Costs",
               "Earned Revenue - JTD", "Actual Cost - JTD Labor", "Actual Hours - JTD",
               "Actual % Complete Estimated - JTD", "Financial Status", "Contract Size Band",
               "Actual Cost - MTD", "Actual Hours - MTD", "Actual Billed - MTD",
               "AP Unapproved Invoice Amt"])
    for num, name, orig, co, billed, actual, projected, estimate, status in jobs:
        ws.append([f"{num} {name}", num, name, orig + co, orig, co, billed,
                   actual, projected, estimate, billed * 0.98, actual * 0.42,
                   round(actual / 160), actual / projected if projected else None,
                   status, "Large" if orig > 5e6 else "Medium",
                   round(actual * 0.055), round(actual / 160 * 0.06), round(billed * 0.04),
                   88412 if num == "26-101" else 0])

    ph = wb.create_sheet("WRH Phase Detail")
    ph.append(["Job and Desc", "Phase and Desc", "Cost Type Desc", "Actual Cost",
               "Current Estimate Costs", "Projected Cost", "Actual Hours - JTD",
               "Remaining Cost - JTD", "Actual Cost - MTD"])
    phase_rows = [
        ("01-100 General", "Labor", 1395880, 1842000, 1861500, 24318, 168420),
        ("01-100 General", "Burden", 558352, 736800, 744600, None, 67368),
        ("54-200 Equipment", "Equipment", 268410, 402500, 396000, 3120, 41220),
        ("92-100 Materials", "Materials", 1043668, 1318400, 1301000, None, 74110),
        ("92-100 Materials", "Subcontract", 512180, 806000, 812500, None, 52700),
        ("54-200 Equipment", "Freight", 71205, 96000, 94300, None, 9062),
        ("01-100 General", "Consumables", 34520, 117000, 108800, None, None),
    ]
    for num, name, *rest in [(j[0], j[1]) for j in jobs]:
        label = f"{num} {name}"
        for phase, ct, act, est, proj, hrs, mtd in phase_rows:
            scale = 1.0 if num == "26-101" else 0.4
            ph.append([label, phase, ct, act * scale, est * scale, proj * scale,
                       hrs, (est - act) * scale, (mtd or 0) * scale or None])

    st = wb.create_sheet("WRH Job Status")
    st.append(["Job and Desc", "Job Status", "Contract Type"])
    for num, name, *_ in jobs:
        st.append([f"{num} {name}", "Open", "Lump Sum"])

    ar = wb.create_sheet("WRH Contract AR")
    ar.append(["Contract", "Contract Status", "Billed Amt", "Received Amt", "Current Retain Amt"])
    for num, name, orig, co, billed, *_ in jobs:
        ar.append([num, "Open", billed, round(billed * 0.9), round(billed * 0.1)])

    meta = wb.create_sheet("Meta")
    meta.append(["as_of", as_of.isoformat()])
    meta.append(["schema_version", 2])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-dir", required=True)
    args = ap.parse_args()
    data_dir = Path(args.data_dir).resolve()
    data_dir.mkdir(parents=True, exist_ok=True)

    workbook = data_dir / "vista" / "Vista Model DEV - Data.xlsx"
    build_workbook(workbook, date.today())

    os.environ["PLANWISE_DATA_DIR"] = str(data_dir)
    os.environ["PLANWISE_VISTA_WORKBOOK"] = str(workbook)

    from backend import auth, briefing, db, lookahead, records, schedule, store, vista

    # Two accounts: the admin the dev signs in with, and a teammate for
    # attribution variety. Idempotent — re-running the seed is safe.
    try:
        auth.bootstrap_admin(auth.setup_token(), "Dev Admin", "planwise-dev")
        # Give the bootstrap admin an email so the login card's email field works.
        conn0 = db.connect()
        conn0.execute("UPDATE users SET email = 'dev@planwise.local' WHERE name = 'Dev Admin'")
        conn0.commit()
    except Exception:
        pass
    try:
        u = auth.register("field@planwise.local", "Terry", "Kowalski", "planwise-dev")
        auth.approve_account(u["name"], actor="Dev Admin")
    except Exception:
        pass

    job = "26-101"
    conn = db.connect()
    if conn.execute("SELECT 1 FROM change_orders WHERE job_number = ?", (job,)).fetchone():
        print("already seeded:", data_dir)
        return 0

    # Contacts + compliance on Job setup.
    store.patch_meta(job, {
        "bond_required": "Yes", "insurance_cert": "Yes",
        "certified_payroll": "No", "pla_davis_bacon": "No",
        "project_manager": "Dev Admin", "superintendent": "J. Benavides",
        "field_leader": "Terry Kowalski", "estimator": "A. Reyes",
        "customer_address": "West Plains Electric Cooperative, 400 Grid Rd",
        "contacts": [
            {"name": "Dana Whitfield", "role": "Owner's representative · WPEC",
             "phone": "806-555-0142", "email": "dwhitfield@wpec.example"},
            {"name": "Miguel Ortega", "role": "Substation engineer · WPEC",
             "phone": "806-555-0188", "email": "mortega@wpec.example"},
        ]}, actor="Dev Admin")

    # Change orders: three approved customer + ONE drafted-unsent (drives the
    # attention item), and two sub COs — one covered by a PO, one exposed.
    for n, desc, amt, status, appr in [
            ("01", "Added duct bank crossing at Rd 1400", 82600, "Approved", 82600),
            ("02", "Relay panel rework — Bay 3", 198750, "Approved", 198750),
            ("03", "Additional grounding per revised E-201", 261560, "Approved", 261560),
            ("04", "Transformer pad anchor revision", 186400, "Unsent", None)]:
        store.add_co(job, {"kind": "customer", "co_number": n, "description": desc,
                           "amount_submitted": amt, "amount_approved": appr,
                           "status": status, "date_submitted": "2026-08-11"},
                     actor="Dev Admin")
    sub_a = store.add_co(job, {"kind": "subcontractor", "co_number": "S-01",
                               "subcontractor": "Caprock Boring",
                               "description": "Bore alignment change", "amount_approved": 60000,
                               "status": "Approved"}, actor="Dev Admin")
    store.add_co(job, {"kind": "subcontractor", "co_number": "S-02",
                       "subcontractor": "Llano Testing LLC",
                       "description": "Second relay test pass", "amount_approved": 36800,
                       "status": "Approved"}, actor="Dev Admin")

    # POs, one issued from sub CO S-01, one exposure (S-02) left uncovered.
    po_rows = [
        ("P26-101-01", "Cinco Steel Supply", "Galvanized structures, Bays 2–4", "Materials", 412600, [("4412", "2026-06-18", 96400), ("4488", "2026-07-16", 84200)]),
        ("P26-101-02", "Caprock Boring", "Duct bank, Rd 1400 crossing", "Subcontract", 286000, [("CB-1190", "2026-06-30", 78000)]),
        ("P26-101-03", "High Plains Crane", "Crane, 90-ton, 6 weeks", "Equipment", 118400, [("HPC-3390", "2026-07-21", 23300)]),
    ]
    for num, vendor, desc, ct, amt, invs in po_rows:
        po = store.add_po(job, {"po_number": num, "vendor": vendor, "description": desc,
                                "cost_type": ct, "original_amount": amt,
                                "order_date": "2026-05-04", "ordered_by": "Dev Admin",
                                "source_co_id": sub_a["id"] if vendor == "Caprock Boring" else None},
                          actor="Dev Admin")
        for inv_num, inv_date, inv_amt in invs:
            store.add_invoice(job, po["id"], {"invoice_number": inv_num, "date": inv_date,
                                              "amount": inv_amt}, actor="Dev Admin")

    # RFIs and submittals: a draft (attention), a sent, an answered-unconfirmed.
    records.add_record(job, "rfi", {"number": "RFI-016", "title": "Control building conduit stub-up count",
                                    "question": "Drawing E-201 calls for conduit stub-ups at the control building but the count differs between plan and schedule. Confirm the required stub-up count and sizes.",
                                    "due_date": "2026-08-28"}, actor="Terry Kowalski")
    sent = records.add_record(job, "rfi", {"number": "RFI-015", "title": "Bus spacing dimension, sheet E-204",
                                           "question": "Sheet E-204 does not dimension bus centerline spacing at the Bay 3 riser. Provide the spacing dimension.",
                                           "due_date": "2026-08-21", "to_name": "Miguel Ortega",
                                           "to_email": "mortega@wpec.example"}, actor="Dev Admin")
    records.save_draft(sent["id"], "RFI-015 · Bus spacing dimension — Job 26-101",
                       "Miguel, sheet E-204 does not dimension the bus centerline spacing at the Bay 3 riser. Please provide the dimension.",
                       source="template", actor="Dev Admin")
    records.mark_sent(sent["id"], actor="Dev Admin")
    records.add_reply(sent["id"], {"from_name": "Miguel Ortega", "from_email": "mortega@wpec.example",
                                   "received_at": "2026-08-18T10:02:00",
                                   "body": "Centerline spacing at the Bay 3 riser is 48 inches. Revised E-204 to follow."})
    records.add_record(job, "submittal", {"number": "SUB-008", "title": "Relay panel shop drawings — Bay 3",
                                          "spec_section": "26 05 26 — Grounding and Bonding",
                                          "status": "Revise & Resubmit", "due_date": "2026-08-19",
                                          "to_name": "Dana Whitfield", "to_email": "dwhitfield@wpec.example"},
                       actor="Dev Admin")

    # Schedule: a summary, tasks with all four link types, one milestone.
    d0 = date(2026, 3, 2)
    def dd(days): return (d0 + timedelta(days=days)).isoformat()
    tasks = [
        ("1", "Northwind Substation — 138kV Upgrade", dd(0), dd(290), 0, 1, ""),
        ("2", "Mobilization & site prep", dd(0), dd(23), 1, 0, ""),
        ("3", "Civil / foundations", dd(24), dd(85), 1, 0, "2FS"),
        ("4", "Underground duct bank", dd(86), dd(135), 1, 0, "3FS"),
        ("5", "Steel & bus assembly", dd(136), dd(190), 1, 0, "4FS"),
        ("6", "Bus & insulator install", dd(150), dd(200), 2, 0, "5SS+14d"),
        ("7", "Relay panels & wiring", dd(191), dd(242), 1, 0, "5FS"),
        ("8", "Commissioning & testing", dd(243), dd(275), 1, 0, "7FS"),
        ("9", "Substation energized", dd(276), dd(276), 1, 1, "8FS"),
        ("10", "Punchlist & closeout", dd(277), dd(290), 1, 0, "9FS"),
    ]
    payload = []
    for ext, name, start, finish, level, ms, preds in tasks:
        payload.append({"external_id": ext, "name": name, "start": start, "finish": finish,
                        "outline_level": level, "is_summary": 1 if ext == "1" else 0,
                        "is_milestone": ms, "percent_complete": 100 if ext in ("2", "3") else 60 if ext == "4" else 0,
                        "predecessors": preds})
    schedule.import_tasks(job, payload, source="manual", mode="replace", actor="Dev Admin")

    # Look ahead + areas.
    la = lookahead.get_or_create(job, lookahead.week_start().isoformat()
                                 if hasattr(lookahead, "week_start") else None, actor="Dev Admin") \
        if hasattr(lookahead, "get_or_create") else None
    for name, color in [("Bay 3", "#1F5F97"), ("Duct bank", "#1B6B3D"), ("Yard", "#7A5100")]:
        try:
            lookahead.add_area(job, {"name": name, "color": color}, actor="Dev Admin")
        except Exception:
            break

    # Briefing seeds itself on first read; force it now so the page has one.
    briefing.get_or_create(job, actor="Dev Admin")

    # Two history points so the forecast chart draws a real line.
    for delta, factor in ((14, 0.94), (0, 1.0)):
        as_of = (date.today() - timedelta(days=delta)).isoformat()
        for num, name, orig, co, billed, actual, projected, estimate, status in [
                ("26-101", "Northwind Substation Upgrade", 5940000, 542910, 4196540, 3884215, 5318900, 5318700, "")]:
            conn.execute(
                "INSERT OR IGNORE INTO vista_history (as_of, job_number, actual_cost,"
                " projected_cost, current_estimate, actual_billed, pct_complete, captured_at)"
                " VALUES (?,?,?,?,?,?,?,?)",
                (as_of, num, actual * factor, projected, estimate, billed * factor,
                 actual * factor / projected, db.now()))
    conn.commit()

    print("seeded:", data_dir)
    print("sign in: dev@planwise.local / planwise-dev")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
