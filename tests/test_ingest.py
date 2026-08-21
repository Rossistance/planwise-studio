"""Vista extract delivery over the API (Phase 5b).

A hosted PlanWise has no OneDrive, so the daily extract is pushed to it. The
property that matters: a bad push must fail for the uploader rather than
quietly break the app for the whole team.
"""
from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient

from backend import app as app_module
from backend import config, db, vista


@pytest.fixture(autouse=True)
def isolate(tmp_path, monkeypatch):
    monkeypatch.setenv("PLANWISE_DATA_DIR", str(tmp_path))
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", "")   # no canonical lookup
    monkeypatch.setenv("PLANWISE_INGEST_TOKEN", "test-ingest-secret")
    db.reset_for_tests()
    vista._cached = None
    yield
    db.reset_for_tests()
    vista._cached = None


@pytest.fixture
def client():
    return TestClient(app_module.app)


def workbook_bytes(*, jobs=(("24-003", "Siemens - Wendell", 5_863_489.0),),
                   with_phases=True, drop_job_column=None) -> bytes:
    """A minimal but genuine extract: the sheets and captions vista.py needs."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = vista.SHEET_JOBS
    job_header = ["Row Labels", "Job Number", "Job Name", "Current Contract Amt",
               "Original Contract Amt", "Change Order Revenue", "Actual Billed",
               "Actual Cost", "Projected Cost", "Current Estimate Costs",
               "Earned Revenue - JTD", "Actual Cost - JTD Labor", "Actual Hours - JTD",
               "Actual % Complete Estimated - JTD", "Financial Status",
               "Contract Size Band"]
    if drop_job_column:
        job_header = [c for c in job_header if c != drop_job_column]
    ws.append(job_header)
    for number, name, contract in jobs:
        ws.append([f"{number} - {name}", number, name, contract, contract, 0,
                   0, 0, 0, 0, 0, 0, 0, 0, "On Track", "Large"])

    phases = wb.create_sheet(vista.SHEET_PHASES)
    phases.append(["Job and Desc", "Phase and Desc", "Cost Type Desc", "Actual Cost",
                   "Current Estimate Costs", "Projected Cost", "Actual Hours - JTD",
                   "Remaining Cost - JTD", "Actual Cost - MTD"])
    if with_phases:
        label = f"{jobs[0][0]} - {jobs[0][1]}"
        phases.append([label, "92-100 - Underground", "Labor",
                       1.0, 2.0, 3.0, 4.0, 5.0, 6.0])

    status = wb.create_sheet(vista.SHEET_STATUS)
    status.append(["Job and Desc", "Job Status", "Contract Type"])
    for number, name, _c in jobs:
        status.append([f"{number} - {name}", "Open", "Lump Sum"])

    ar = wb.create_sheet(vista.SHEET_CONTRACT_AR)
    ar.append(["Contract", "Contract Status", "Billed Amt", "Received Amt",
               "Current Retain Amt"])
    for number, _n, _c in jobs:
        ar.append([number, "1-Open", 0.0, 0.0, 0.0])

    meta = wb.create_sheet(vista.SHEET_META)
    meta.append(["key", "value"])
    meta.append(["as_of", "2026-08-09T06:30:52-04:00"])
    meta.append(["schema_version", "2"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def push(client, data: bytes, token="test-ingest-secret", name="Vista Model 2026 - Data.xlsx"):
    headers = {"X-PlanWise-Ingest": token} if token is not None else {}
    return client.post("/api/vista/workbook", headers=headers,
                       files={"file": (name, data,
                                       "application/vnd.openxmlformats-officedocument."
                                       "spreadsheetml.sheet")})


# --- the happy path -----------------------------------------------------------

def test_a_pushed_workbook_becomes_the_one_the_app_reads():
    from fastapi.testclient import TestClient

    c = TestClient(app_module.app)
    assert config.vista_workbook() is None          # nothing to read yet

    r = push(c, workbook_bytes())
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["jobs"] == 1
    assert body["phase_rows"] == 1
    assert str(body["schema_version"]) == "2"
    assert body["received_bytes"] > 0

    assert config.vista_workbook() == config.pushed_workbook()
    assert vista.load().jobs["24-003"]["job_name"] == "Siemens - Wendell"


def test_the_push_is_reflected_immediately_not_on_the_next_cache_miss(client):
    push(client, workbook_bytes(jobs=(("24-003", "First Name", 1.0),)))
    assert vista.load().jobs["24-003"]["job_name"] == "First Name"

    push(client, workbook_bytes(jobs=(("24-003", "Second Name", 2.0),)))
    assert vista.load().jobs["24-003"]["job_name"] == "Second Name"


def test_health_reports_the_pushed_workbook(client):
    assert client.get("/api/health").json()["workbook_found"] is False
    push(client, workbook_bytes())
    health = client.get("/api/health").json()
    assert health["workbook_found"] is True
    assert health["vista"]["ok"] is True
    assert health["vista"]["job_count"] == 1


# --- refusing bad pushes ------------------------------------------------------

def test_a_workbook_that_does_not_parse_is_refused_and_the_old_one_survives(client):
    """The property this whole endpoint exists to guarantee."""
    push(client, workbook_bytes(jobs=(("24-003", "Good Data", 99.0),)))
    good = config.pushed_workbook().read_bytes()

    r = push(client, b"this is not a spreadsheet at all")
    assert r.status_code == 422
    assert "did not parse" in r.json()["detail"]

    # untouched, still readable, still the good data
    assert config.pushed_workbook().read_bytes() == good
    assert vista.load().jobs["24-003"]["job_name"] == "Good Data"


def test_a_schema_drifted_workbook_is_refused_and_names_what_is_wrong(client):
    """Every shape of bad workbook comes back 422 with the reason, not a 500."""
    push(client, workbook_bytes())

    # a required column quietly disappears from an otherwise valid extract
    r = push(client, workbook_bytes(drop_job_column="Actual Cost"))
    assert r.status_code == 422
    assert "actual cost" in r.json()["detail"].lower()

    # a whole sheet missing
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = vista.SHEET_JOBS
    buf = io.BytesIO()
    wb.save(buf)
    r = push(client, buf.getvalue())
    assert r.status_code == 422
    assert vista.SHEET_PHASES.lower() in r.json()["detail"].lower()

    assert vista.load().jobs                              # previous data intact


def test_the_previous_copy_is_kept_for_recovery(client):
    push(client, workbook_bytes(jobs=(("24-003", "Monday", 1.0),)))
    push(client, workbook_bytes(jobs=(("24-003", "Tuesday", 2.0),)))
    prev = config.pushed_workbook_previous()
    assert prev.is_file()

    import openpyxl
    rolled_back = openpyxl.load_workbook(prev, read_only=True)[vista.SHEET_JOBS]
    assert "Monday" in str(list(rolled_back.iter_rows(values_only=True))[1])


def test_a_failed_push_never_becomes_the_file_the_app_reads(client):
    """The scratch file may linger on Windows — openpyxl can still hold its
    handle. What must never happen is a half-written or unparseable file
    becoming the live workbook."""
    push(client, b"garbage")
    assert not config.pushed_workbook().exists()
    assert config.vista_workbook() is None

    push(client, workbook_bytes(jobs=(("24-003", "Good", 1.0),)))
    good = config.pushed_workbook().read_bytes()
    push(client, b"garbage again")
    assert config.pushed_workbook().read_bytes() == good


# --- the token ----------------------------------------------------------------

def test_the_wrong_token_no_token_and_an_unconfigured_server_are_all_refused(client, monkeypatch):
    assert push(client, workbook_bytes(), token="wrong-secret").status_code == 401
    assert push(client, workbook_bytes(), token=None).status_code == 401
    assert config.vista_workbook() is None            # nothing was written

    # Unset means OFF, not open: a misconfigured server refuses uploads
    # rather than accepting anonymous ones.
    monkeypatch.delenv("PLANWISE_INGEST_TOKEN")
    r = push(client, workbook_bytes())
    assert r.status_code == 503
    assert "not configured" in r.json()["detail"]


def test_ingest_does_not_require_a_signed_in_session(client):
    """It runs from an unattended scheduled task — there is no human to sign
    in. The ingest token is the whole credential."""
    assert client.get("/api/users").status_code == 401       # session-gated
    assert push(client, workbook_bytes()).status_code == 200  # token-gated, no session


# --- the companion's own credential (regression, 2026-08-10) -------------------

def test_the_companion_can_file_replies_without_a_session(client, monkeypatch):
    """Phase 5a's session gate locked the companion out of the very endpoint
    its whole job depends on, and the poller reported the 401 as 'no replies
    found'. It runs with no browser open, so it has no session — its own token
    has to be enough for exactly these writes."""
    from backend import ai

    token = ai.companion_token()
    assert token

    paths = ["/api/records/some-id/replies", "/api/records/some-id/sent"]
    for path in paths:
        # no credential at all -> still refused
        assert client.post(path, json={}).status_code == 401
        # wrong token -> refused
        assert client.post(path, json={},
                           headers={"X-PlanWise-Companion": "not-the-token"}).status_code == 401
        # right token -> gets PAST the gate (404 = no such record, which is
        # the endpoint answering rather than the middleware refusing)
        r = client.post(path, json={}, headers={"X-PlanWise-Companion": token})
        assert r.status_code != 401, f"{path} still gated"


def test_the_companion_token_is_not_a_skeleton_key(client):
    """It buys the two calls a companion makes, and nothing else."""
    from backend import ai

    headers = {"X-PlanWise-Companion": ai.companion_token()}
    for path in ["/api/users", "/api/jobs/24-003", "/api/outbox", "/api/companion/token"]:
        assert client.get(path, headers=headers).status_code == 401, f"{path} was reachable"


# --- the settings-triggered refresh (2.0.3) -----------------------------------
# The server only HOLDS a refresh request; the pull runs on the one PC with
# the Power BI connection, which learns about it through the companion poll.

def _signed_in():
    from backend import auth
    c = TestClient(app_module.app)
    auth.bootstrap_admin(auth.setup_token(), "Ross Hixon", "a-good-password")
    assert c.post("/api/auth/login",
                  json={"name": "Ross Hixon", "password": "a-good-password"}).status_code == 200
    return c


def test_a_refresh_request_is_held_shown_on_health_and_rides_the_poll():
    from backend import ai
    c = _signed_in()

    r = c.post("/api/vista/refresh-request")
    assert r.status_code == 200
    assert r.json()["requested_by"] == "Ross Hixon"

    hv = c.get("/api/health").json()["vista"]
    assert hv["refresh_requested_at"] is not None
    assert hv["refresh_requested_by"] == "Ross Hixon"

    manifest = c.get("/api/companion/poll",
                     params={"token": ai.companion_token()}).json()
    assert manifest["vista"]["wanted"] is True


def test_a_refresh_request_needs_a_session():
    c = TestClient(app_module.app)
    assert c.post("/api/vista/refresh-request").status_code == 401


def test_a_workbook_push_clears_the_standing_request():
    c = _signed_in()
    assert c.post("/api/vista/refresh-request").status_code == 200

    r = c.post("/api/vista/workbook",
               files={"file": ("Vista Model 2026 - Data.xlsx", io.BytesIO(workbook_bytes()),
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
               headers={"X-PlanWise-Ingest": "test-ingest-secret"})
    assert r.status_code == 200, r.text

    hv = c.get("/api/health").json()["vista"]
    assert hv["refresh_requested_at"] is None


def test_an_unserved_request_expires_rather_than_burning_forever():
    from datetime import datetime, timedelta, timezone
    c = _signed_in()
    assert c.post("/api/vista/refresh-request").status_code == 200

    stale = (datetime.now(timezone.utc) - timedelta(minutes=45)).isoformat(timespec="seconds")
    conn = db.connect()
    conn.execute("UPDATE settings SET value = ? WHERE key = 'vista_refresh_requested_at'", (stale,))
    conn.commit()

    hv = c.get("/api/health").json()["vista"]
    assert hv["refresh_requested_at"] is None
