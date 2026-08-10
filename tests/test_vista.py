"""Vista reader tests.

These build a synthetic workbook in tmp_path rather than reading the real
Company Share file: tests must pass on a machine that has never synced it, and
must never depend on today's live numbers.
"""
from __future__ import annotations

import openpyxl
import pytest

from backend import vista


@pytest.fixture(autouse=True)
def isolate(tmp_path, monkeypatch):
    monkeypatch.setenv("PLANWISE_DATA_DIR", str(tmp_path / "data"))
    vista._cached = None
    yield
    vista._cached = None


def build_workbook(path, *, jobs=None, phases=None, as_of="2026-08-08T07:17:23-04:00",
                   v2=False, v2_job_extra=None, contracts=None):
    """v2=False builds a schema-v1 workbook (no MTD/AP columns, no AR sheet) —
    what a machine with a stale OneDrive copy still has."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = vista.SHEET_JOBS
    headers = list(vista.JOB_COLUMNS) + (list(vista.JOB_COLUMNS_V2) if v2 else [])
    ws.append(headers)
    for row in jobs if jobs is not None else [DEFAULT_JOB]:
        ws.append(row + (v2_job_extra or [None] * 4 if v2 else []))

    ph = wb.create_sheet(vista.SHEET_PHASES)
    ph.append(list(vista.PHASE_COLUMNS) + (list(vista.PHASE_COLUMNS_V2) if v2 else []))
    for row in phases if phases is not None else DEFAULT_PHASES:
        ph.append(row)

    st = wb.create_sheet(vista.SHEET_STATUS)
    st.append(list(vista.STATUS_COLUMNS))
    st.append(["24-003 - Siemens - Wendell", "1-Open", "Lump Sum"])

    if v2:
        ar = wb.create_sheet(vista.SHEET_CONTRACT_AR)
        ar.append(["Contract", "Contract Status", "Contract Amt",
                   "Billed Amt", "Received Amt", "Current Retain Amt"])
        for row in contracts if contracts is not None else [DEFAULT_CONTRACT]:
            ar.append(row)

    meta = wb.create_sheet(vista.SHEET_META)
    meta.append(["Key", "Value"])
    if as_of:
        meta.append(["as_of", as_of])
    meta.append(["schema_version", "2" if v2 else "1"])

    wb.save(path)
    return path


# Column order matches JOB_COLUMNS declaration order.
DEFAULT_JOB = [
    "24-003 - Siemens - Wendell", "24-003", "Siemens - Wendell",
    5863489, 5707533, 155956, 5859364, 4897552.64, 4912163.32, 5386866.81,
    None, 125877.8, 2951.5, 0.909165348381799, "On Track", "$5M-$10M",
]

DEFAULT_PHASES = [
    # job, phase, cost type, actual, estimate, projected, hours, remaining
    ["24-003 - Siemens - Wendell", "95-100 - EPC", "Labor", 125141.56, 134800.8, 134800.8, 2933.5, -125141.56],
    ["24-003 - Siemens - Wendell", "92-600 - Supplemental Insurances", "Labor", 736.24, 0, 0, 18, -736.24],
    ["24-003 - Siemens - Wendell", "95-100 - EPC", "Equipment", 165064.89, 67379, 55379, None, -165064.89],
]

# Contract, Status, Contract Amt, Billed, Received, Retainage — the real
# 24-003 values, which match the Vista Crystal export to the penny.
DEFAULT_CONTRACT = ["24-003", "1-Open", 5863489, 5859364, 5273427.62, 585936.38]


@pytest.fixture
def workbook(tmp_path, monkeypatch):
    p = build_workbook(tmp_path / "vista.xlsx")
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(p))
    return p


def test_loads_jobs_phases_and_as_of(workbook):
    snap = vista.load()
    assert len(snap.jobs) == 1
    assert snap.jobs["24-003"]["job_name"] == "Siemens - Wendell"
    assert snap.jobs["24-003"]["current_contract"] == 5863489
    assert snap.as_of is not None
    assert len(snap.phases["24-003 - Siemens - Wendell"]) == 3


def test_job_status_is_folded_onto_the_job(workbook):
    snap = vista.load()
    assert snap.jobs["24-003"]["job_status"] == "1-Open"
    assert snap.jobs["24-003"]["contract_type"] == "Lump Sum"


def test_blank_is_none_not_zero(workbook):
    """The 2.0 line's most expensive bug: a blank cell rendered as $0.00."""
    snap = vista.load()
    assert snap.jobs["24-003"]["earned_revenue"] is None
    equipment = next(r for r in snap.phases["24-003 - Siemens - Wendell"]
                     if r["cost_type"] == "Equipment")
    assert equipment["hours_units"] is None
    # ...while a real zero survives as a real zero.
    insurance = next(r for r in snap.phases["24-003 - Siemens - Wendell"]
                     if r["phase"].startswith("92-600"))
    assert insurance["current_estimate"] == 0


def test_cost_types_come_from_the_data(workbook):
    """No fixed taxonomy: this job has two cost types, so two rows."""
    types = vista.cost_types_for(vista.load(), "24-003")
    assert [t["cost_type"] for t in types] == ["Equipment", "Labor"]
    labor = next(t for t in types if t["cost_type"] == "Labor")
    assert labor["phase_count"] == 2
    assert labor["actual_cost"] == pytest.approx(125877.80)
    assert labor["current_estimate"] == pytest.approx(134800.80)


def test_unknown_cost_type_is_kept_not_dropped(tmp_path, monkeypatch):
    p = build_workbook(tmp_path / "v.xlsx", phases=[
        ["24-003 - Siemens - Wendell", "95-100 - EPC", "Bespoke Rigging", 10.0, 20.0, 15.0, None, 10.0],
    ])
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(p))
    types = vista.cost_types_for(vista.load(), "24-003")
    assert [t["cost_type"] for t in types] == ["Bespoke Rigging"]


def test_variance_is_estimate_minus_actual(tmp_path, monkeypatch):
    """Not estimate-minus-projected. Vista's projection can sit below actual on
    a stale line, which made an overspent line report a surplus."""
    p = build_workbook(tmp_path / "v.xlsx", phases=[
        # est 67,379 / actual 165,065 / projected 55,379 — the real 24-003
        # Equipment line. estimate - projected would read +12,000 (green).
        ["24-003 - Siemens - Wendell", "95-100 - EPC", "Equipment", 165064.89, 67379.0, 55379.0, None, None],
    ])
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(p))
    equip = vista.cost_types_for(vista.load(), "24-003")[0]
    assert equip["variance"] == pytest.approx(-97685.89)


def test_variance_needs_an_estimate_and_an_actual(tmp_path, monkeypatch):
    p = build_workbook(tmp_path / "v.xlsx", phases=[
        ["24-003 - Siemens - Wendell", "95-100 - EPC", "Labor", 100.0, None, 200.0, None, None],
    ])
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(p))
    assert vista.cost_types_for(vista.load(), "24-003")[0]["variance"] is None


def test_pct_complete_is_none_without_a_projection(tmp_path, monkeypatch):
    p = build_workbook(tmp_path / "v.xlsx", phases=[
        ["24-003 - Siemens - Wendell", "95-100 - EPC", "Labor", 100.0, 200.0, None, None, None],
    ])
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(p))
    assert vista.cost_types_for(vista.load(), "24-003")[0]["pct_complete"] is None


def test_missing_workbook_is_an_honest_error(monkeypatch):
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", "")
    with pytest.raises(vista.VistaUnavailable, match="not found"):
        vista.load()


def test_wrong_shape_names_the_missing_sheet(tmp_path, monkeypatch):
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(tmp_path / "wrong.xlsx")
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(tmp_path / "wrong.xlsx"))
    with pytest.raises(vista.VistaUnavailable, match="Pivot Data"):
        vista.load()


def test_schema_drift_names_the_missing_column(tmp_path, monkeypatch):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = vista.SHEET_JOBS
    ws.append(["Row Labels", "Job Number"])  # the rest went away
    wb.create_sheet(vista.SHEET_PHASES).append(list(vista.PHASE_COLUMNS))
    wb.create_sheet(vista.SHEET_STATUS).append(list(vista.STATUS_COLUMNS))
    wb.save(tmp_path / "drift.xlsx")
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(tmp_path / "drift.xlsx"))
    with pytest.raises(vista.VistaUnavailable, match="Current Contract Amt"):
        vista.load()


def test_empty_extract_is_refused(tmp_path, monkeypatch):
    p = build_workbook(tmp_path / "empty.xlsx", jobs=[], phases=[])
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(p))
    with pytest.raises(vista.VistaUnavailable, match="zero jobs"):
        vista.load()


def test_stale_when_as_of_is_old(tmp_path, monkeypatch):
    p = build_workbook(tmp_path / "old.xlsx", as_of="2026-07-01T03:37:00-04:00")
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(p))
    assert vista.load().is_stale is True


def test_v2_workbook_loads_mtd_ap_and_contract_ar(tmp_path, monkeypatch):
    p = build_workbook(tmp_path / "v2.xlsx", v2=True,
                       v2_job_extra=[1510.30, 12.5, None, 250.0])
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(p))
    snap = vista.load()
    assert snap.schema_version == 2
    j = snap.jobs["24-003"]
    assert j["mtd_cost"] == pytest.approx(1510.30)
    assert j["mtd_hours"] == pytest.approx(12.5)
    assert j["mtd_billed"] is None          # blank stays blank in v2 too
    assert j["unapproved_ap"] == pytest.approx(250.0)
    ar = snap.contract_ar["24-003"]
    assert ar["billed"] == pytest.approx(5859364.00)
    assert ar["collected"] == pytest.approx(5273427.62)
    assert ar["retainage"] == pytest.approx(585936.38)


def test_v1_workbook_still_loads_with_v2_fields_none(workbook):
    """A machine whose OneDrive hasn't synced the regenerated file must not
    crash — v2 fields read as not-reported, never as zero."""
    snap = vista.load()
    assert snap.schema_version == 1
    assert snap.contract_ar == {}
    j = snap.jobs["24-003"]
    for f in ("mtd_cost", "mtd_hours", "mtd_billed", "unapproved_ap"):
        assert j[f] is None


def test_job_numbers_sort_naturally(tmp_path, monkeypatch):
    rows = []
    for num in ["10", "2", "1", "24-003", "2-017"]:
        r = list(DEFAULT_JOB)
        r[0], r[1] = f"{num} - J{num}", num
        rows.append(r)
    p = build_workbook(tmp_path / "sort.xlsx", jobs=rows)
    monkeypatch.setenv("PLANWISE_VISTA_WORKBOOK", str(p))
    assert vista.job_numbers(vista.load()) == ["1", "2", "2-017", "10", "24-003"]
